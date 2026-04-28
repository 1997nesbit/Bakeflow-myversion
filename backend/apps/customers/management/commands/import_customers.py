"""
Management command: import_customers

Reads 'contacts- cleaned.xlsx' (POS_Export sheet) and bulk-inserts records
into the Customer model.

Rules:
  - Only name + phone are imported (email/notes optional if available)
  - Phone is normalised to local Tanzanian format (07XXXXXXXX)
  - Rows missing a valid phone are skipped
  - Names with obvious anomalies (emoji, Instagram handles, date suffixes,
    pure symbols, very short/long, non-ASCII-dominant) get replaced with
    "Customer <N>" placeholders
  - Uses get_or_create(phone=...) — safe to re-run, never duplicates

Usage:
    python manage.py import_customers
    python manage.py import_customers --file /path/to/other.xlsx
    python manage.py import_customers --dry-run
"""

import re
import unicodedata
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.customers.models import Customer

# ---------------------------------------------------------------------------
# Defaults
# ---------------------------------------------------------------------------
def _default_xlsx() -> Path:
    """Locate the contacts xlsx relative to Django's BASE_DIR (backend/)."""
    try:
        from django.conf import settings
        # BASE_DIR = backend/ → .parent = Bakeflow---OG/ → .parent = bakery/
        return Path(settings.BASE_DIR).parent.parent / "contacts- cleaned.xlsx"
    except Exception:
        # Fallback: resolve relative to this file (7 parents = bakery/)
        return Path(__file__).resolve().parents[6].parent / "contacts- cleaned.xlsx"

DEFAULT_XLSX = _default_xlsx()
SHEET_NAME = "POS_Export"

# Column indices in POS_Export (0-based)
COL_FIRST  = 0
COL_MIDDLE = 1
COL_LAST   = 2
COL_PHONE1 = 3
COL_PHONE2 = 4
COL_PHONE3 = 5
COL_EMAIL1 = 6
COL_NOTES  = 9


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Regex: strip trailing date-tag appended by the phone owner  e.g. "Aug2025", "28Aug 2023"
_DATE_SUFFIX = re.compile(
    r"""
    [\s\-_]*        # optional separator
    (?:
        \d{1,2}     # optional leading day
        [\s\-]?
    )?
    (?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)
    [\s\-]?
    \d{2,4}         # year
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# Tanzanian mobile prefixes (after stripping country code)
_TZ_PREFIX = re.compile(r"^0[16789]\d{8}$")


_TZ_SHORT = re.compile(r"^[6789]\d{8}$")    # 9-digit with no leading 0 (valid TZ prefixes: 6,7,8,9)


def _normalise_single(raw: str) -> str | None:
    """
    Normalise one phone candidate to '07XXXXXXXX' local format.

    Handles all common Tanzanian formats:
      +255XXXXXXXXX  →  strip country code  → 0XXXXXXXXX
       255XXXXXXXXX  →  strip country code  → 0XXXXXXXXX
      0255XXXXXXXXX  →  strip country code  → 0XXXXXXXXX
       0XXXXXXXXX   →  already local format
        XXXXXXXXX   →  9-digit, prepend 0   → 0XXXXXXXXX  (e.g. 785857396)

    Returns None for anything else.
    """
    digits = re.sub(r"[\s\-().+]", "", raw)

    # Country-code variants → strip to local 10-digit
    if digits.startswith("255") and len(digits) == 12:
        digits = "0" + digits[3:]
    elif digits.startswith("0255") and len(digits) == 13:
        digits = "0" + digits[4:]
    # 9-digit bare number (leading 0 was omitted): 7XX XXX XXX, 6XX XXX XXX …
    elif _TZ_SHORT.match(digits):
        digits = "0" + digits

    if _TZ_PREFIX.match(digits):
        return digits
    return None


def normalise_phone(raw: str) -> str | None:
    """
    Normalise a raw phone string to local format '07XXXXXXXX'.

    Handles cells that contain multiple numbers joined with ' ::: '
    (Google Contacts multi-value export separator) — returns the first
    valid Tanzanian number found in the cell.

    Returns None if no valid number is found.
    """
    if not raw:
        return None
    # Split on the Google Contacts multi-value separator before anything else
    candidates = str(raw).split(":::")
    for candidate in candidates:
        result = _normalise_single(candidate.strip())
        if result:
            return result
    return None


def _is_anomalous_name(name: str) -> bool:
    """
    Return True if the name looks like an alias, handle, or garbage string
    rather than a real person's name.
    """
    if not name or len(name.strip()) < 2:
        return True

    s = name.strip()

    # Contains emoji or lots of non-ASCII
    non_ascii = sum(1 for c in s if ord(c) > 127)
    if non_ascii > len(s) * 0.4:          # >40 % non-ASCII chars
        return True

    # Starts with _, __, @, #, *, +
    if re.match(r"^[_@#*+.!?]", s):
        return True

    # Contains Instagram/handle patterns
    if re.search(r"[._]{2,}", s):          # __ or ..
        return True

    # Looks like a URL or IG reference
    if re.search(r"https?://|www\.|\.com|IG$|ig$", s):
        return True

    # Pure digits or mostly digits
    digit_ratio = sum(c.isdigit() for c in s) / max(len(s), 1)
    if digit_ratio > 0.5:
        return True

    # After stripping the date suffix, fewer than 2 real chars remain
    stripped = _DATE_SUFFIX.sub("", s).strip()
    if len(stripped) < 2:
        return True

    return False


def clean_name(raw_first, raw_middle, raw_last) -> str | None:
    """
    Combine name parts, strip date suffixes, and return a cleaned name.
    Returns None if the result is anomalous (caller will use placeholder).
    """
    parts = [p for p in (raw_first, raw_middle, raw_last) if p and str(p).strip()]
    combined = " ".join(str(p).strip() for p in parts)
    # Strip date suffix from the combined name
    combined = _DATE_SUFFIX.sub("", combined).strip()
    # Normalise unicode (NFKC collapses fancy chars)
    combined = unicodedata.normalize("NFKC", combined).strip()

    if _is_anomalous_name(combined):
        return None
    return combined if combined else None


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = "Import customers from the BBR contacts Excel file."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=str(DEFAULT_XLSX),
            help="Path to the contacts xlsx file (default: auto-detected)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Print what would be imported without writing to the database.",
        )
        parser.add_argument(
            "--no-log",
            action="store_true",
            help="Suppress writing the failures log file.",
        )

    def handle(self, *args, **options):
        xlsx_path  = Path(options["file"])
        dry_run    = options["dry_run"]
        write_log  = not options["no_log"]
        log_path   = xlsx_path.parent / "import_failures.txt"

        if not xlsx_path.exists():
            raise CommandError(f"File not found: {xlsx_path}")

        self.stdout.write(f"Opening: {xlsx_path}")
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(
                f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[SHEET_NAME]

        counters = {
            "total":            0,
            "created":          0,
            "skipped_phone":    0,
            "skipped_duplicate": 0,
            "placeholder":      0,
        }
        failed_rows = []   # (row_number, raw_name, raw_phones) for every skipped row

        # Placeholder counter: base on existing DB count so numbers stay stable on re-runs
        placeholder_base = Customer.objects.count() if not dry_run else 0
        placeholder_n    = placeholder_base

        header_skipped = False
        row_number     = 1   # header is row 1

        for row in ws.iter_rows(values_only=True):
            # Skip the header row
            if not header_skipped:
                header_skipped = True
                continue

            row_number     += 1
            counters["total"] += 1

            # ── Phone ──────────────────────────────────────────────────────
            phone      = None
            raw_phones = []
            for col in (COL_PHONE1, COL_PHONE2, COL_PHONE3):
                if col < len(row) and row[col]:
                    raw_phones.append(str(row[col]))
                    if not phone:
                        phone = normalise_phone(str(row[col]))

            if not phone:
                counters["skipped_phone"] += 1
                # Capture raw name for the failure log
                raw_name_parts = [
                    str(row[c]) if len(row) > c and row[c] else ""
                    for c in (COL_FIRST, COL_MIDDLE, COL_LAST)
                ]
                raw_name = " ".join(p for p in raw_name_parts if p).strip() or "(no name)"
                failed_rows.append((row_number, raw_name, raw_phones))
                continue

            # ── Name ───────────────────────────────────────────────────────
            name = clean_name(
                row[COL_FIRST]  if len(row) > COL_FIRST  else None,
                row[COL_MIDDLE] if len(row) > COL_MIDDLE else None,
                row[COL_LAST]   if len(row) > COL_LAST   else None,
            )
            if name is None:
                placeholder_n += 1
                name = f"Customer {placeholder_n}"
                counters["placeholder"] += 1

            # ── Email (optional) ───────────────────────────────────────────
            email = None
            if len(row) > COL_EMAIL1 and row[COL_EMAIL1]:
                raw_email = str(row[COL_EMAIL1]).strip()
                if "@" in raw_email:
                    email = raw_email

            # ── Notes (optional) ───────────────────────────────────────────
            notes = ""
            if len(row) > COL_NOTES and row[COL_NOTES]:
                notes = str(row[COL_NOTES]).strip()

            # ── Dry-run or insert ──────────────────────────────────────────
            if dry_run:
                safe_name = name.encode("ascii", errors="backslashreplace").decode("ascii")
                self.stdout.write(f"  [DRY] {safe_name!r:40s}  {phone}  {email or ''}")
                counters["created"] += 1
                continue

            obj, created = Customer.objects.get_or_create(
                phone=phone,
                defaults=dict(name=name, email=email, notes=notes),
            )
            if created:
                counters["created"] += 1
            else:
                counters["skipped_duplicate"] += 1

        wb.close()

        # ── Write failures log ─────────────────────────────────────────────
        if write_log and failed_rows:
            with open(log_path, "w", encoding="utf-8") as lf:
                lf.write(f"Import failures — {xlsx_path.name}\n")
                lf.write("Rows skipped because no valid Tanzanian phone could be parsed.\n")
                lf.write(f"Total: {len(failed_rows)}\n")
                lf.write("=" * 80 + "\n\n")
                for rn, name, phones in failed_rows:
                    raw = " | ".join(phones) if phones else "(no phone columns)"
                    # Encode to ASCII so the log file is always clean
                    safe_name = name.encode("ascii", errors="backslashreplace").decode("ascii")
                    lf.write(f"Row {rn:>6}  Name: {safe_name:<40s}  Raw phone(s): {raw}\n")

        # ── Summary ────────────────────────────────────────────────────────
        label = "[DRY RUN] " if dry_run else ""
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(f"{label}Done!"))
        self.stdout.write(f"  Total rows read     : {counters['total']}")
        self.stdout.write(self.style.SUCCESS(
            f"  Created             : {counters['created']}"
        ))
        self.stdout.write(self.style.WARNING(
            f"  Name placeholders   : {counters['placeholder']}"
        ))
        self.stdout.write(f"  Skipped (no phone)  : {counters['skipped_phone']}")
        if not dry_run:
            self.stdout.write(f"  Skipped (duplicate) : {counters['skipped_duplicate']}")
        if write_log and failed_rows:
            self.stdout.write(self.style.WARNING(
                f"  Failures log        : {log_path}"
            ))
